#!/usr/bin/env python3
"""
AppDynamics API Integration Tool
A comprehensive script to retrieve application performance metrics from AppDynamics
and export them to CSV or Excel format.

Features:
- OAuth2 authentication with Client ID/Secret
- Retrieves application performance metrics (CPM, Response Time, Errors, Health)
- Supports both CSV and Excel export with formatting
- Handles token management and API errors gracefully
- Progress indication and user-friendly interface

Author: AppDynamics Integration Team
Version: 1.0
"""

import os
import sys
import json
import time
import logging
import argparse
import threading
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Tuple, Any
import urllib.parse

# Required imports - install with: pip install requests pandas openpyxl
try:
    import requests
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
except ImportError as e:
    print(f"❌ Missing required dependency: {e}")
    print("📦 Install required packages: pip install requests pandas openpyxl")
    sys.exit(1)

# Configuration - Modify these values or set environment variables
CONFIG = {
    # AppDynamics Controller Configuration
    'CONTROLLER_URL': os.getenv('APPDYNAMICS_CONTROLLER_URL', 'https://bny-ucf.saas.appdynamics.com'),
    'CLIENT_ID': os.getenv('APPDYNAMICS_CLIENT_ID', 'your-client-id'),
    'CLIENT_SECRET': os.getenv('APPDYNAMICS_CLIENT_SECRET', 'your-client-secret'),
    'ACCOUNT_NAME': os.getenv('APPDYNAMICS_ACCOUNT_NAME', 'bny-ucf'),
    
    # API Configuration
    'API_TIMEOUT': 30,
    'MAX_RETRIES': 3,
    'BACKOFF_FACTOR': 1,
    
    # Data Collection Settings
    'DEFAULT_TIME_RANGE_HOURS': 1,
    'MAX_APPLICATIONS': 200,
    'THREAD_COUNT': 5,
    
    # Output Settings
    'DEFAULT_CSV_FILENAME': f'appdynamics_metrics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv',
    'DEFAULT_EXCEL_FILENAME': f'appdynamics_metrics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
}

# Global variables for token management
token_data = {
    'access_token': None,
    'expires_at': None,
    'lock': threading.Lock()
}

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class AppDynamicsAPI:
    """
    AppDynamics REST API client with OAuth2 authentication and metrics retrieval capabilities.
    """
    
    def __init__(self, controller_url: str, client_id: str, client_secret: str, account_name: str):
        """
        Initialize the AppDynamics API client.
        
        Args:
            controller_url: AppDynamics controller URL
            client_id: OAuth2 client ID
            client_secret: OAuth2 client secret
            account_name: AppDynamics account name
        """
        self.controller_url = controller_url.rstrip('/')
        self.client_id = client_id
        self.client_secret = client_secret
        self.account_name = account_name
        
        # Setup session with retry strategy
        self.session = requests.Session()
        retry_strategy = Retry(
            total=CONFIG['MAX_RETRIES'],
            backoff_factor=CONFIG['BACKOFF_FACTOR'],
            status_forcelist=[429, 500, 502, 503, 504]
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)
        
        self.session.headers.update({
            'User-Agent': 'AppDynamics-Metrics-Tool/1.0',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        })
    
    def authenticate(self) -> bool:
        """
        Authenticate with AppDynamics using OAuth2 client credentials flow.
        
        Returns:
            True if authentication successful, False otherwise
        """
        with token_data['lock']:
            # Check if we have a valid token
            if self._is_token_valid():
                return True
            
            try:
                print("🔐 Authenticating with AppDynamics...")
                
                # Prepare OAuth2 request
                auth_url = f"{self.controller_url}/controller/api/oauth/access_token"
                
                payload = {
                    'grant_type': 'client_credentials',
                    'client_id': self.client_id,
                    'client_secret': self.client_secret
                }
                
                headers = {
                    'Content-Type': 'application/vnd.appd.cntrl+protobuf;v=1',
                    'Accept': 'application/vnd.appd.cntrl+protobuf;v=1'
                }
                
                response = self.session.post(
                    auth_url,
                    data=payload,
                    headers=headers,
                    timeout=CONFIG['API_TIMEOUT']
                )
                
                if response.status_code == 200:
                    auth_data = response.json()
                    token_data['access_token'] = auth_data['access_token']
                    expires_in = auth_data.get('expires_in', 3600)
                    token_data['expires_at'] = datetime.now() + timedelta(seconds=expires_in - 300)  # 5min buffer
                    
                    self.session.headers.update({
                        'Authorization': f"Bearer {token_data['access_token']}"
                    })
                    
                    print("✅ Authentication successful")
                    return True
                else:
                    logger.error(f"Authentication failed: {response.status_code} - {response.text}")
                    return False
                    
            except Exception as e:
                logger.error(f"Authentication error: {e}")
                return False
    
    def _is_token_valid(self) -> bool:
        """Check if current token is valid and not expired."""
        return (token_data['access_token'] is not None and 
                token_data['expires_at'] is not None and 
                datetime.now() < token_data['expires_at'])
    
    def get_applications(self) -> List[Dict[str, Any]]:
        """
        Retrieve list of all applications from AppDynamics.
        
        Returns:
            List of application dictionaries with id and name
        """
        try:
            print("📋 Retrieving applications list...")
            
            url = f"{self.controller_url}/controller/rest/applications"
            params = {
                'output': 'json'
            }
            
            response = self.session.get(url, params=params, timeout=CONFIG['API_TIMEOUT'])
            
            if response.status_code == 200:
                applications = response.json()
                print(f"✅ Found {len(applications)} applications")
                return applications
            else:
                logger.error(f"Failed to retrieve applications: {response.status_code} - {response.text}")
                return []
                
        except Exception as e:
            logger.error(f"Error retrieving applications: {e}")
            return []
    
    def get_application_metrics(self, app_id: int, app_name: str, 
                              start_time: datetime, end_time: datetime) -> Dict[str, Any]:
        """
        Retrieve performance metrics for a specific application.
        
        Args:
            app_id: Application ID
            app_name: Application name
            start_time: Start time for metrics
            end_time: End time for metrics
            
        Returns:
            Dictionary containing application metrics
        """
        metrics_data = {
            'application_name': app_name,
            'application_id': app_id,
            'calls_per_minute': 0,
            'average_response_time': 0,
            'errors_per_minute': 0,
            'error_percentage': 0,
            'health_status': 'Unknown',
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        try:
            # Convert times to AppD format (milliseconds since epoch)
            start_time_ms = int(start_time.timestamp() * 1000)
            end_time_ms = int(end_time.timestamp() * 1000)
            
            # Define metric paths to retrieve
            metric_paths = [
                'Application Infrastructure Performance|*|Calls per Minute',
                'Application Infrastructure Performance|*|Average Response Time (ms)',
                'Application Infrastructure Performance|*|Errors per Minute'
            ]
            
            base_url = f"{self.controller_url}/controller/rest/applications/{app_id}/metric-data"
            
            for metric_path in metric_paths:
                try:
                    params = {
                        'metric-path': metric_path,
                        'time-range-type': 'BETWEEN_TIMES',
                        'start-time': start_time_ms,
                        'end-time': end_time_ms,
                        'rollup': 'true',
                        'output': 'json'
                    }
                    
                    response = self.session.get(base_url, params=params, timeout=CONFIG['API_TIMEOUT'])
                    
                    if response.status_code == 200:
                        metric_data = response.json()
                        self._process_metric_data(metric_data, metric_path, metrics_data)
                    
                except Exception as e:
                    logger.warning(f"Failed to retrieve {metric_path} for {app_name}: {e}")
                    continue
            
            # Calculate derived metrics
            if metrics_data['calls_per_minute'] > 0:
                metrics_data['error_percentage'] = round(
                    (metrics_data['errors_per_minute'] / metrics_data['calls_per_minute']) * 100, 2
                )
            
            # Determine health status
            metrics_data['health_status'] = self._determine_health_status(metrics_data)
            
        except Exception as e:
            logger.error(f"Error retrieving metrics for {app_name}: {e}")
        
        return metrics_data
    
    def _process_metric_data(self, metric_data: List[Dict], metric_path: str, 
                           result: Dict[str, Any]) -> None:
        """Process and extract metric values from AppD response."""
        if not metric_data:
            return
            
        try:
            # Get the latest metric value
            for metric in metric_data:
                metric_values = metric.get('metricValues', [])
                if metric_values:
                    latest_value = metric_values[-1].get('value', 0)
                    
                    if 'Calls per Minute' in metric_path:
                        result['calls_per_minute'] = round(float(latest_value), 2)
                    elif 'Average Response Time' in metric_path:
                        result['average_response_time'] = round(float(latest_value), 2)
                    elif 'Errors per Minute' in metric_path:
                        result['errors_per_minute'] = round(float(latest_value), 2)
                        
        except Exception as e:
            logger.warning(f"Error processing metric data for {metric_path}: {e}")
    
    def _determine_health_status(self, metrics: Dict[str, Any]) -> str:
        """
        Determine application health status based on metrics.
        
        Args:
            metrics: Application metrics dictionary
            
        Returns:
            Health status string (Healthy, Warning, Critical, Unknown)
        """
        try:
            error_rate = metrics['error_percentage']
            response_time = metrics['average_response_time']
            
            # Define health thresholds
            if error_rate > 5 or response_time > 5000:
                return 'Critical'
            elif error_rate > 1 or response_time > 2000:
                return 'Warning'
            elif metrics['calls_per_minute'] > 0:
                return 'Healthy'
            else:
                return 'No Traffic'
                
        except Exception:
            return 'Unknown'


class MetricsExporter:
    """
    Handles exporting metrics data to CSV or Excel format with proper formatting.
    """
    
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str) -> bool:
        """
        Export metrics data to CSV format.
        
        Args:
            data: List of metrics dictionaries
            filename: Output CSV filename
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            print(f"📊 Exporting data to CSV: {filename}")
            
            df = pd.DataFrame(data)
            
            # Reorder columns for better readability
            column_order = [
                'application_name', 'calls_per_minute', 'average_response_time', 
                'errors_per_minute', 'error_percentage', 'health_status', 'timestamp'
            ]
            df = df.reindex(columns=column_order)
            
            # Rename columns for display
            df.columns = [
                'Application Name', 'Calls/Min', 'Avg Response Time (ms)', 
                'Errors/Min', 'Error Rate (%)', 'Health Status', 'Timestamp'
            ]
            
            df.to_csv(filename, index=False, encoding='utf-8')
            print(f"✅ CSV export completed: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return False
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], filename: str) -> bool:
        """
        Export metrics data to Excel format with formatting.
        
        Args:
            data: List of metrics dictionaries
            filename: Output Excel filename
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            print(f"📊 Exporting data to Excel: {filename}")
            
            df = pd.DataFrame(data)
            
            # Reorder columns
            column_order = [
                'application_name', 'calls_per_minute', 'average_response_time', 
                'errors_per_minute', 'error_percentage', 'health_status', 'timestamp'
            ]
            df = df.reindex(columns=column_order)
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "AppDynamics Metrics"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center")
            
            # Health status colors
            health_colors = {
                'Healthy': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                'Warning': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                'Critical': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            }
            
            # Add headers
            headers = [
                'Application Name', 'Calls/Min', 'Avg Response Time (ms)', 
                'Errors/Min', 'Error Rate (%)', 'Health Status', 'Timestamp'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Add data rows
            for row_idx, row_data in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=row_data['application_name'])
                ws.cell(row=row_idx, column=2, value=row_data['calls_per_minute'])
                ws.cell(row=row_idx, column=3, value=row_data['average_response_time'])
                ws.cell(row=row_idx, column=4, value=row_data['errors_per_minute'])
                ws.cell(row=row_idx, column=5, value=row_data['error_percentage'])
                
                # Health status with color coding
                health_cell = ws.cell(row=row_idx, column=6, value=row_data['health_status'])
                health_status = row_data['health_status']
                if health_status in health_colors:
                    health_cell.fill = health_colors[health_status]
                
                ws.cell(row=row_idx, column=7, value=row_data['timestamp'])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add footer with timestamp
            footer_row = len(data) + 3
            ws.cell(row=footer_row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            wb.save(filename)
            print(f"✅ Excel export completed: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return False


def get_time_range(hours: int = None) -> Tuple[datetime, datetime]:
    """
    Get start and end time for metrics collection.
    
    Args:
        hours: Number of hours to go back (if None, prompt user)
        
    Returns:
        Tuple of (start_time, end_time)
    """
    if hours is None:
        print("\n⏰ Time Range Selection:")
        print("1. Last hour")
        print("2. Last 24 hours")
        print("3. Custom range")
        
        while True:
            choice = input("Select time range (1-3): ").strip()
            if choice == '1':
                hours = 1
                break
            elif choice == '2':
                hours = 24
                break
            elif choice == '3':
                try:
                    hours = int(input("Enter hours to go back: "))
                    if hours > 0:
                        break
                    else:
                        print("❌ Please enter a positive number")
                except ValueError:
                    print("❌ Please enter a valid number")
            else:
                print("❌ Please select 1, 2, or 3")
    
    end_time = datetime.now()
    start_time = end_time - timedelta(hours=hours)
    
    print(f"📅 Time range: {start_time.strftime('%Y-%m-%d %H:%M')} to {end_time.strftime('%Y-%m-%d %H:%M')}")
    return start_time, end_time


def get_export_format() -> str:
    """
    Get user preference for export format.
    
    Returns:
        Export format ('csv' or 'excel')
    """
    print("\n📁 Export Format Selection:")
    print("1. CSV (Comma-separated values)")
    print("2. Excel (Formatted spreadsheet)")
    
    while True:
        choice = input("Select export format (1-2): ").strip()
        if choice == '1':
            return 'csv'
        elif choice == '2':
            return 'excel'
        else:
            print("❌ Please select 1 or 2")


def get_output_filename(export_format: str) -> str:
    """
    Get output filename from user or use default.
    
    Args:
        export_format: Export format ('csv' or 'excel')
        
    Returns:
        Output filename
    """
    default_filename = (CONFIG['DEFAULT_CSV_FILENAME'] if export_format == 'csv' 
                       else CONFIG['DEFAULT_EXCEL_FILENAME'])
    
    filename = input(f"\n📝 Enter filename (or press Enter for '{default_filename}'): ").strip()
    
    if not filename:
        filename = default_filename
    
    # Ensure correct extension
    expected_ext = '.csv' if export_format == 'csv' else '.xlsx'
    if not filename.lower().endswith(expected_ext):
        filename += expected_ext
    
    return filename


def display_summary_stats(data: List[Dict[str, Any]]) -> None:
    """
    Display summary statistics in console.
    
    Args:
        data: List of metrics dictionaries
    """
    if not data:
        return
    
    print("\n📊 Summary Statistics:")
    print("=" * 50)
    
    total_apps = len(data)
    healthy = sum(1 for app in data if app['health_status'] == 'Healthy')
    warning = sum(1 for app in data if app['health_status'] == 'Warning')
    critical = sum(1 for app in data if app['health_status'] == 'Critical')
    
    total_calls = sum(app['calls_per_minute'] for app in data)
    total_errors = sum(app['errors_per_minute'] for app in data)
    avg_response_time = sum(app['average_response_time'] for app in data) / total_apps if total_apps > 0 else 0
    
    print(f"Total Applications: {total_apps}")
    print(f"Healthy: {healthy} | Warning: {warning} | Critical: {critical}")
    print(f"Total Calls/Min: {total_calls:,.2f}")
    print(f"Total Errors/Min: {total_errors:,.2f}")
    print(f"Average Response Time: {avg_response_time:.2f}ms")
    
    if total_calls > 0:
        overall_error_rate = (total_errors / total_calls) * 100
        print(f"Overall Error Rate: {overall_error_rate:.2f}%")


def collect_metrics_threaded(api: AppDynamicsAPI, applications: List[Dict], 
                           start_time: datetime, end_time: datetime) -> List[Dict[str, Any]]:
    """
    Collect metrics for all applications using threading for better performance.
    
    Args:
        api: AppDynamics API instance
        applications: List of applications
        start_time: Start time for metrics
        end_time: End time for metrics
        
    Returns:
        List of metrics dictionaries
    """
    metrics_data = []
    completed = 0
    total = len(applications)
    lock = threading.Lock()
    
    def collect_app_metrics(app):
        nonlocal completed
        try:
            app_metrics = api.get_application_metrics(
                app['id'], app['name'], start_time, end_time
            )
            with lock:
                metrics_data.append(app_metrics)
                completed += 1
                if completed % 5 == 0 or completed == total:
                    progress = (completed / total) * 100
                    print(f"📊 Progress: {completed}/{total} applications ({progress:.1f}%)")
        except Exception as e:
            logger.error(f"Failed to collect metrics for {app['name']}: {e}")
            with lock:
                completed += 1
    
    print(f"🔄 Collecting metrics for {total} applications...")
    
    # Create and start threads
    threads = []
    for app in applications:
        thread = threading.Thread(target=collect_app_metrics, args=(app,))
        threads.append(thread)
        thread.start()
        
        # Limit concurrent threads
        if len(threads) >= CONFIG['THREAD_COUNT']:
            for t in threads:
                t.join()
            threads.clear()
    
    # Wait for remaining threads
    for thread in threads:
        thread.join()
    
    print(f"✅ Metrics collection completed: {len(metrics_data)} applications processed")
    return metrics_data


def main():
    """
    Main function to orchestrate the AppDynamics metrics collection and export process.
    """
    print("🚀 AppDynamics API Integration Tool")
    print("=" * 50)
    
    # Validate configuration
    if (CONFIG['CLIENT_ID'] == 'your-client-id' or 
        CONFIG['CLIENT_SECRET'] == 'your-client-secret'):
        print("❌ Error: Please configure your AppDynamics credentials")
        print("Set environment variables:")
        print("- APPDYNAMICS_CLIENT_ID")
        print("- APPDYNAMICS_CLIENT_SECRET") 
        print("- APPDYNAMICS_CONTROLLER_URL (optional)")
        print("- APPDYNAMICS_ACCOUNT_NAME (optional)")
        return 1
    
    try:
        # Initialize API client
        api = AppDynamicsAPI(
            CONFIG['CONTROLLER_URL'],
            CONFIG['CLIENT_ID'],
            CONFIG['CLIENT_SECRET'],
            CONFIG['ACCOUNT_NAME']
        )
        
        # Authenticate
        if not api.authenticate():
            print("❌ Authentication failed. Please check your credentials.")
            return 1
        
        # Get applications
        applications = api.get_applications()
        if not applications:
            print("❌ No applications found or failed to retrieve applications.")
            return 1
        
        # Get time range
        start_time, end_time = get_time_range()
        
        # Collect metrics
        metrics_data = collect_metrics_threaded(api, applications, start_time, end_time)
        
        if not metrics_data:
            print("❌ No metrics data collected.")
            return 1
        
        # Display summary
        display_summary_stats(metrics_data)
        
        # Get export preferences
        export_format = get_export_format()
        filename = get_output_filename(export_format)
        
        # Export data
        exporter = MetricsExporter()
        
        if export_format == 'csv':
            success = exporter.export_to_csv(metrics_data, filename)
        else:
            success = exporter.export_to_excel(metrics_data, filename)
        
        if success:
            print(f"\n🎉 Export completed successfully!")
            print(f"📄 File: {filename}")
            print(f"📊 Records: {len(metrics_data)}")
            return 0
        else:
            print("❌ Export failed.")
            return 1
            
    except KeyboardInterrupt:
        print("\n⚠️ Operation cancelled by user.")
        return 1
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    # Setup command line arguments
    parser = argparse.ArgumentParser(description='AppDynamics Metrics Collection Tool')
    parser.add_argument('--hours', type=int, help='Hours to look back for metrics (default: interactive)')
    parser.add_argument('--format', choices=['csv', 'excel'], help='Export format (default: interactive)')
    parser.add_argument('--output', help='Output filename (default: auto-generated)')
    parser.add_argument('--verbose', action='store_true', help='Enable verbose logging')
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Override interactive prompts with command line arguments if provided
    if args.hours:
        get_time_range = lambda h=None: get_time_range(args.hours)
    if args.format:
        get_export_format = lambda: args.format
    if args.output:
        get_output_filename = lambda fmt: args.output
    
    sys.exit(main())